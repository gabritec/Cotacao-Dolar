# MODULOS NECESSÁRIO
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import load_workbook




# PEGANDO A COTAÇÃO DO DOLÁR E ARMAZENANDO EM UMA VÁRIAVEL
navegador = webdriver.Chrome('C:Caminho do chromedriver.exe ') # Entre parenteses informe o caminho do executavel chromedriver 
navegador.get("https://www.google.com/") # Vai estar entrando no google
navegador.find_element(By.XPATH,
    'Codigo Xpath da barra de pesquisa do google').send_keys("cotação dolar") # Através do codigo "Xpath" da barra de pesquisa do google ele pesquisará "cotação dolar"
navegador.find_element(By.XPATH,
    'Codigo Xpath da barra de pesquisa do google').send_keys(Keys.ENTER) # Através do codigo "Xpath" da barra de pesquisa do google ele irá apertar enter para efetuar a pesquisa
cotacao_dolar = navegador.find_element(By.XPATH,
    'Codigo Xpath do valor da cotação do dolar').get_attribute("data-value") # Estará pegando o Valor do "Xpath" informado e armazenando em uma variavável


# FORMATAÇÃO
cotacao_float = float(cotacao_dolar) # Transformando a variáve que está em str em float
cotacao_dolar2= ('{:.2f}'.format(cotacao_float)) # Transformando em apenas 2 casas decimais formatando

print(cotacao_dolar2) # Caso precise vizualizar antes de transferir para algum lugar



# Armazenando a "cotação_dolar2" dentro de uma celula de uma planilha do excel
wb = load_workbook(r'C: Caminho do arquivo desejado.xlsx')      # Está carregando uma planilha em xlsx e armazenando em uma váriavel
ws = wb['Plan1']            # Selecionando a planilha('mostra na parte inferior direito')
cell = ws['N3']             # Selecione a célula
cell.value = float(cotacao_dolar2)          # A célula está recebendo o valor da variável que está armazenado a cotação
cell.number_format = '0.00'      # Formatando em forma de números(obs: Se não colocar esses detalhes qualquer formula de calculo sobre a célula não irá funcionar.)
wb.save(f'C:Caminho do arquivo desejado.xlsx')      # Salvando na mesma pasta em que você carregou ela, pois dessa forma ela irá subtituir automaticamente
